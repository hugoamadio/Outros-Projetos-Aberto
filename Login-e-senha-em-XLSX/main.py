import tkinter as tk
import openpyxl
import bcrypt
import re
from tkinter import messagebox

def cadastrar():    #função cadastrar
    usuario = entradaUsuario.get()  #recebe usuario
    senha = entradaSenha.get()      #recebe senha

    if not re.search(r'[A-Z]', senha):      #verifica se a senha tem pelo menos uma letra maiuscula
        print("A Senha deve conter pelo menos uma letra MAIÚSCULA.")
        return

    criptografia = bcrypt.hashpw(senha.encode("utf-8"), bcrypt.gensalt())   #criptografa a senha

    try:        #abre a planilha
        arquivoxlsx = openpyxl.load_workbook("LS.xlsx")
    except FileNotFoundError:
        print("Database não encontrado")

    planilha = arquivoxlsx.active   #usa a sheet1

    linha = [usuario, criptografia]
    planilha.append(linha)
    arquivoxlsx.save("LS.xlsx")
    print("Usuário e Senha Cadastrados: ", usuario, senha)

def login():    #função login
    usuario = entradaUsuario.get()
    senha = entradaSenha.get()

    try:
        arquivoxlsx = openpyxl.load_workbook("LS.xlsx")
    except FileNotFoundError:
        print("Database não encontrado")
        return

    planilha = arquivoxlsx.active

    for linha in planilha.iter_rows(values_only=True):  #verifica quais linhas retornam valores nas colunas A e B
        if len(linha) >= 2 and linha[0] == usuario:
            if bcrypt.checkpw(senha.encode('utf-8'), linha[1].encode('utf-8')):
                messagebox.showinfo("Login", "Login Aprovado")
                return
    messagebox.showinfo("Inválido", "Usuário ou Senha incorreta / Não Cadastrado")


root = tk.Tk()  #cria a janela do tkinter
root.title("Tela de Login")
root.configure(background="lightblue")
root.geometry("300x200")

# widgets de usuario e senha
textoUsuario = tk.Label(root, text="Usuário:", background="lightblue")
textoUsuario.grid(row=1, column=1, padx=10, pady=5, sticky="w")
entradaUsuario = tk.Entry(root)
entradaUsuario.grid(row=1, column=2, padx=10, pady=5)

textoSenha = tk.Label(root, text="Senha:", background="lightblue")
textoSenha.grid(row=2, column=1, padx=10, pady=5, sticky="w")
entradaSenha = tk.Entry(root, show="*")
entradaSenha.grid(row=2, column=2, padx=10, pady=5)

# botões
botaoLogin = tk.Button(root, text="Login", command=login)
botaoLogin.grid(row=3, column=1, padx=10, pady=10)
botaoCadastrar = tk.Button(root, text="Cadastrar", command=cadastrar)
botaoCadastrar.grid(row=3, column=2, padx=10, pady=10)


root.mainloop()